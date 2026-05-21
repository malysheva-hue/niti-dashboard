#!/usr/bin/env python3
"""Сборка дашборда «Нити Творчества» — Этап 1.

Читает:
  data/csv-year/*.csv          — историческая база (12 месяцев)
  inputs/*.xlsx                — свежак (план-факт мая, дневные воронки, шаблон цен)
  data/daily_summary.md        — ежедневная сводка от РОПа

Пишет:
  data.json                    — общий лёгкий срез (KPI, агрегаты, OOS, цены, сводка)
  data/processed/months/*.json — полный срез SKU за каждый месяц
  index.html                   — встраивает копию data.json в маркер /* APP_DATA_PLACEHOLDER */
"""
from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent
CSV_DIR = ROOT / "data" / "csv-year"
INPUTS_DIR = ROOT / "inputs"
PROCESSED_DIR = ROOT / "data" / "processed"
MONTHS_DIR = PROCESSED_DIR / "months"
DAILY_SUMMARY_MD = ROOT / "data" / "daily_summary.md"
INDEX_HTML = ROOT / "index.html"
DATA_JSON = ROOT / "data.json"

INCOMPLETE_MONTHS = {"2025-06", "2026-05"}
NO_MANAGERS_MONTHS = {"2026-02"}
DEFAULT_MONTH = "2026-05"

NAME_NORMALIZE = {
    "Цигельникова Виктория": "Виктория",
}

# v2: Антон — РОП, его SKU перешли к Владимиру (передача дел с 06.2025).
# Везде на дашборде «Антон» как менеджер → «Владимир».
MANAGER_REMAP = {"Антон": "Владимир"}

def remap_mgr(name):
    if not name:
        return name
    name = NAME_NORMALIZE.get(name, name)
    return MANAGER_REMAP.get(name, name)

STAR_CODES = {"КЛ007", "КЛ003", "ЛК013", "ПВС002"}
LOCO_RISK_CODES = {"ПС002", "ФМПНН0032", "ЮТ001", "ЮТ002"}
ACTIVE_MANAGERS = {"Виктория", "Настя", "Владимир"}
AI_SUMMARY_FILE = ROOT / "data" / "ai_summary.json"
DAILY_REVENUE_FILE = CSV_DIR / "Дневная_Сумма_заказов.csv"   # v2.0: legacy, не используется
DAILY_PROFIT_FILE = CSV_DIR / "Дневная_Прибыль.csv"          # v2.0: legacy, не используется
HISTORY_PARQUET = ROOT / "data" / "history.parquet"           # v2.0: годовая база


def _load_history_pivot(indicators: list[str]) -> dict:
    """v2.0: читает data/history.parquet, возвращает {indicator: {code: {date: value}}}.
    Фильтрует только нужные показатели для экономии памяти.
    """
    if not HISTORY_PARQUET.exists():
        return {}
    try:
        import pyarrow.parquet as pq
        import pyarrow.compute as pc
    except ImportError:
        return {}
    try:
        t = pq.read_table(
            HISTORY_PARQUET,
            columns=["code", "indicator", "date", "value"],
            filters=[("indicator", "in", indicators)],
        )
    except Exception:
        t = pq.read_table(HISTORY_PARQUET,
                          columns=["code", "indicator", "date", "value"])
        mask = pc.is_in(t["indicator"], options=pc.SetLookupOptions(value_set=indicators))
        t = t.filter(mask)
    out = {}
    codes = t["code"].to_pylist()
    inds = t["indicator"].to_pylist()
    dates = t["date"].to_pylist()
    values = t["value"].to_pylist()
    for c, i, d, v in zip(codes, inds, dates, values):
        if v is None or c is None or i is None or d is None:
            continue
        out.setdefault(i, {}).setdefault(c, {})[str(d)] = float(v)
    return out


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "Не указан план", "—"):
        return None
    s = s.replace(",", ".").replace("\xa0", "").replace(" ", "")
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def round_n(v, ndigits=2):
    if v is None:
        return None
    return round(v, ndigits)


def _date_key(s):
    """Ключ для корректной сортировки строк формата dd.mm.yyyy.
    Возвращает (yyyy, mm, dd). Битые/пустые строки → (0,0,0)."""
    if not s:
        return (0, 0, 0)
    parts = str(s).split(".")
    if len(parts) != 3:
        return (0, 0, 0)
    try:
        d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
        return (y, m, d)
    except (TypeError, ValueError):
        return (0, 0, 0)


def parse_summary_csv():
    out = []
    with open(CSV_DIR / "00_Сводка_по_месяцам.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            m = row.get("Месяц")
            if not m:
                continue
            out.append({
                "month": m,
                "revenue": num(row.get("Сумма заказов, руб")),
                "orders": num(row.get("Факт заказов, шт")),
                "profit": num(row.get("Прибыль, руб")),
                "ad_spend": num(row.get("Рекламные расходы, руб")),
                "frozen": num(row.get("Замороженный капитал")),
                "margin": num(row.get("Маржа %")),
                "drr": num(row.get("ДРР %")),
                "incomplete": m in INCOMPLETE_MONTHS,
            })
    return out


def parse_managers_csv():
    out = defaultdict(dict)
    field_map = [
        ("Сумма заказов, руб", "revenue"),
        ("Факт заказов, шт", "orders"),
        ("Прибыль, руб", "profit"),
        ("Рекламные расходы, руб", "ad_spend"),
        ("Замороженный капитал", "frozen"),
    ]
    with open(CSV_DIR / "01_Менеджеры_по_месяцам.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            m = row.get("Месяц")
            mgr = row.get("Менеджер")
            if not m or not mgr:
                continue
            mgr = remap_mgr(mgr)
            acc = out[m].setdefault(mgr, {k: 0.0 for _, k in field_map})
            for src, dst in field_map:
                v = num(row.get(src))
                if v is not None:
                    acc[dst] = (acc[dst] or 0) + v
    # пересчёт взвешенной маржи после агрегации
    for m_key, mgrs in out.items():
        for mgr_key, d in mgrs.items():
            d["margin"] = (d["profit"] / d["revenue"] * 100) if d.get("revenue") else None
    return dict(out)


def parse_categories_csv():
    out = defaultdict(list)
    with open(CSV_DIR / "03_Категории_по_месяцам.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            m = row.get("Месяц")
            if not m:
                continue
            out[m].append({
                "name": row.get("Категория"),
                "revenue": num(row.get("Сумма заказов, руб")),
                "orders": num(row.get("Факт заказов, шт")),
                "profit": num(row.get("Прибыль, руб")),
                "ad_spend": num(row.get("Рекламные расходы, руб")),
                "frozen": num(row.get("Замороженный капитал")),
            })
    return dict(out)


def parse_sku_year_csv():
    out = []
    with open(CSV_DIR / "04_SKU_итоги_года.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            out.append({
                "code": row.get("Артикул поставщика"),
                "sku": row.get("SKU"),
                "name": row.get("Название товара"),
                "category": row.get("Категория"),
                "revenue": num(row.get("Сумма заказов, руб")),
                "orders": num(row.get("Факт заказов, шт")),
                "profit": num(row.get("Прибыль, руб")),
                "ad_spend": num(row.get("Рекламные расходы, руб")),
                "frozen": num(row.get("Замороженный капитал")),
                "margin": num(row.get("Маржа %")),
                "drr": num(row.get("ДРР %")),
            })
    return out


def parse_sku_monthly_csv():
    out = {}
    for path in sorted(CSV_DIR.glob("SKU_*.csv")):
        m = re.search(r"SKU_(\d{4}-\d{2})\.csv", path.name)
        if not m:
            continue
        month = m.group(1)
        rows = []
        with open(path, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                mgr = row.get("Менеджер") or ""
                mgr = remap_mgr(mgr)
                rows.append({
                    "code": row.get("Артикул поставщика"),
                    "sku": row.get("SKU"),
                    "name": row.get("Название товара"),
                    "category": row.get("Категория"),
                    "group": row.get("Группа товаров") or "",
                    "manager": mgr,
                    "revenue": num(row.get("Сумма заказов, руб")),
                    "orders": num(row.get("Факт заказов, шт")),
                    "profit": num(row.get("Прибыль, руб")),
                    "margin": num(row.get("Маржинальность")),
                    "ad_spend": num(row.get("Рекламные расходы, руб")),
                    "drr": num(row.get("ДРР от рекламных продаж")),
                    "tacos": num(row.get("ДРР от продаж")),
                    "stock": num(row.get("Остаток")),
                    "turnover": num(row.get("Оборачиваемость, дней")),
                    "frozen": num(row.get("Замороженный капитал")),
                })
        out[month] = rows
    return out


def parse_plan_fact_xlsx():
    """Самый свежий План_факт_*.xlsx → {code: {indicator: {totals, day_str: value}}}."""
    import openpyxl
    candidates = sorted(INPUTS_DIR.glob("План_факт_*.xlsx"),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        return {}, None
    src = candidates[0]
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb["План-Факт"]

    header = None
    data = defaultdict(dict)
    current_code = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = row
            continue
        code = row[2]
        if code:
            current_code = code
        if not current_code:
            continue
        indicator = row[7]
        if not indicator:
            continue
        row_data = {"totals": row[8]}
        for c in range(9, len(row)):
            day_label = header[c] if c < len(header) else None
            if day_label:
                row_data[str(day_label)] = row[c]
        data[current_code][indicator] = row_data

    wb.close()
    return dict(data), src.name


def parse_funnels_xlsx():
    """Дневные воронки → {
        'by_code': {date: {code: {shows,ctr,clicks,carts}}}  # ТОЛЬКО для звёзд и топ-локо
        'overall': {date: {shows, clicks, carts, ctr_avg, orders_amount, buyouts_amount, buyout_pct}}
    }."""
    import openpyxl
    targeted = STAR_CODES | LOCO_RISK_CODES
    by_code = {}
    overall = {}
    for path in sorted(p for p in INPUTS_DIR.glob("*.xlsx") if "воронка" in p.name.lower()):
        m = re.search(r"с (\d+)-(\d+)-(\d+) по", path.name)
        if not m:
            continue
        day_iso = f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Товары"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            wb.close()
            continue
        header = rows[1]
        idx = {h: i for i, h in enumerate(header) if h}
        day_data = {}
        sum_shows = sum_clicks = sum_carts = 0
        ctr_values = []
        for r in rows[2:]:
            code = r[idx["Артикул продавца"]] if "Артикул продавца" in idx else None
            if not code:
                continue
            shows = num(r[idx["Показы"]]) if "Показы" in idx else None
            clicks = num(r[idx["Переходы в карточку"]]) if "Переходы в карточку" in idx else None
            carts = num(r[idx["Положили в корзину"]]) if "Положили в корзину" in idx else None
            ctr = num(r[idx["CTR"]]) if "CTR" in idx else None
            if shows: sum_shows += shows
            if clicks: sum_clicks += clicks
            if carts: sum_carts += carts
            if ctr is not None: ctr_values.append(ctr)
            if code in targeted:
                day_data[code] = {"shows": shows, "ctr": ctr, "clicks": clicks, "carts": carts}
        by_code[day_iso] = day_data

        # v2.3.1: лист «Фильтры» даёт суточные итоги выкупов и заказов
        orders_amount = buyouts_amount = buyout_pct = None
        try:
            ws2 = wb["Фильтры"]
            rr2 = list(ws2.iter_rows(values_only=True))
            if len(rr2) >= 3:
                h2 = rr2[1]
                vals = rr2[2]
                ix = {h: i for i, h in enumerate(h2) if h}
                def gv(name):
                    i = ix.get(name)
                    if i is None: return None
                    v = vals[i]
                    return float(v) if isinstance(v, (int, float)) else None
                orders_amount = gv("Заказали на сумму, ₽")
                buyouts_amount = gv("Выкупили на сумму, ₽")
                buyout_pct = gv("Процент выкупа")
        except KeyError:
            pass

        overall[day_iso] = {
            "shows": int(sum_shows),
            "clicks": int(sum_clicks),
            "carts": int(sum_carts),
            "ctr_avg": round_n(sum(ctr_values)/len(ctr_values), 2) if ctr_values else None,
            "orders_amount": round_n(orders_amount, 0) if orders_amount is not None else None,
            "buyouts_amount": round_n(buyouts_amount, 0) if buyouts_amount is not None else None,
            "buyout_pct": round_n(buyout_pct, 2) if buyout_pct is not None else None,
        }
        wb.close()
    return {"by_code": by_code, "overall": overall}


def parse_prices_xlsx():
    """Шаблон цен → {code: {price, discount, stock, turnover}}."""
    import openpyxl
    candidates = sorted(INPUTS_DIR.glob("Шаблон обновления цен*.xlsx"),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        return {}, None
    src = candidates[0]
    wb = openpyxl.load_workbook(src, data_only=True)
    # имя листа меняется от выгрузки к выгрузке — берём первый
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        return {}, src.name
    header = rows[0]
    idx = {h: i for i, h in enumerate(header) if h}
    out = {}
    for r in rows[1:]:
        code = r[idx["Артикул продавца"]] if "Артикул продавца" in idx else None
        if not code:
            continue
        s_wb = num(r[idx["Остатки WB"]]) if "Остатки WB" in idx else None
        s_seller = num(r[idx["Остатки продавца"]]) if "Остатки продавца" in idx else None
        stock_total = None
        if s_wb is not None or s_seller is not None:
            stock_total = (s_wb or 0) + (s_seller or 0)
        price_base = num(r[idx["Текущая цена"]]) if "Текущая цена" in idx else None
        discount_seller = num(r[idx["Текущая скидка"]]) if "Текущая скидка" in idx else None
        price_with_discount = num(r[idx["Цена со скидкой"]]) if "Цена со скидкой" in idx else None
        # v1.9: «Наша цена» = базовая × (1 − скидка/100). Если xlsx даёт готовую — берём её.
        price_after_seller = price_with_discount
        if price_after_seller is None and price_base is not None and discount_seller is not None:
            price_after_seller = price_base * (1 - discount_seller / 100)
        out[code] = {
            # legacy keys (для совместимости с уже встроенным HTML)
            "price": price_base,
            "discount": discount_seller,
            "price_with_discount": price_with_discount,
            "stock_wb": s_wb,
            "stock_seller": s_seller,
            "stock_total": stock_total,
            "turnover": num(r[idx["Оборачиваемость"]]) if "Оборачиваемость" in idx else None,
            # v1.9 — явные «наши» поля
            "price_seller_base": price_base,
            "price_seller_discount": discount_seller,
            "price_after_seller": round_n(price_after_seller, 0) if price_after_seller is not None else None,
        }
    wb.close()
    return out, src.name


def _sum_indicator_in_month(pivot, indicator, month):
    """Сумма значений индикатора по дням месяца. Возвращает (total, n_days).

    Значения в parquet — дневная разноска (план/факт «в этот день»).
    """
    y, mm = month.split("-")
    suffix = f".{mm}.{y}"
    days = {}
    for code_map in (pivot.get(indicator) or {}).values():
        for d, v in code_map.items():
            if not str(d).endswith(suffix):
                continue
            days[d] = days.get(d, 0) + (v or 0)
    return (sum(days.values()) if days else 0.0, len(days))


def _days_in_month(month):
    from datetime import date as _date
    y, mm = month.split("-")
    y, mm = int(y), int(mm)
    return (_date(y if mm < 12 else y + 1, (mm % 12) + 1, 1) - _date(y, mm, 1)).days


def compute_month_plans(month):
    """v2.3.1 БЛОК 1: план месяца через parquet, экстраполированный на полный месяц.

    "Плановая выручка" — дневная разноска (план/день). Сумма за 19 дней × 31/19 = месячный план.
    Возвращает {month: {revenue, orders_qty, sales_qty, order_amount, profit, margin_pct, source}}.
    """
    pivot = _load_history_pivot([
        "Плановая выручка", "Плановая сумма заказов, руб",
        "План продаж, шт", "План заказов, шт",
        "Плановая прибыль в день, руб",
    ])
    if not pivot:
        return {}
    dim = _days_in_month(month)

    def total(ind):
        s, n = _sum_indicator_in_month(pivot, ind, month)
        if n == 0:
            return None
        return s * dim / n   # экстраполяция на полный месяц

    plan_rev = total("Плановая выручка")
    plan_order_amount = total("Плановая сумма заказов, руб")
    plan_sales_qty = total("План продаж, шт")
    plan_orders_qty = total("План заказов, шт")
    plan_profit = total("Плановая прибыль в день, руб")
    if not any(v is not None for v in (plan_rev, plan_order_amount, plan_sales_qty, plan_orders_qty, plan_profit)):
        return {}
    margin_pct = (plan_profit / plan_rev * 100) if (plan_rev and plan_profit) else None
    return {
        month: {
            "revenue":      round_n(plan_rev, 0) if plan_rev else None,
            "order_amount": round_n(plan_order_amount, 0) if plan_order_amount else None,
            "sales_qty":    round_n(plan_sales_qty, 0) if plan_sales_qty else None,
            "orders_qty":   round_n(plan_orders_qty, 0) if plan_orders_qty else None,
            "profit":       round_n(plan_profit, 0) if plan_profit else None,
            "margin_pct":   round_n(margin_pct, 2) if margin_pct is not None else None,
            "source":       "history.parquet (extrapolated × dim/days_with_data)",
        }
    }


def compute_month_buyouts(month):
    """v2.3.1 БЛОК 2: факт выручки (по выкупам) и сумма заказов за месяц из parquet.

    "Выручка, руб" — дневная разноска по выкупам. Сумма за все дни месяца = факт месяца к этой дате.
    """
    pivot = _load_history_pivot(["Выручка, руб", "Выручка по заказам, руб"])
    if not pivot:
        return {}
    rev_buyouts, n1 = _sum_indicator_in_month(pivot, "Выручка, руб", month)
    rev_orders, n2 = _sum_indicator_in_month(pivot, "Выручка по заказам, руб", month)
    if n1 == 0 and n2 == 0:
        return {}
    pct = (rev_buyouts / rev_orders * 100) if (rev_orders and rev_buyouts) else None
    return {
        month: {
            "buyouts_amount": round_n(rev_buyouts, 0) if rev_buyouts else None,
            "orders_amount":  round_n(rev_orders, 0) if rev_orders else None,
            "buyout_pct":     round_n(pct, 2) if pct is not None else None,
            "days_with_data": max(n1, n2),
        }
    }


def compute_sku_plans(plan_fact):
    """v2.3 БЛОК 5: достаём планы по SKU из План-Факта.

    Возвращает {code: {plan_revenue, plan_orders, plan_profit, fact_revenue, fact_orders, fact_profit}}.
    Только SKU с реальными планами.
    """
    PLAN_MAP = {
        "Плановая сумма заказов, руб": "plan_revenue",
        "План заказов, шт": "plan_orders",
        "Прогноз прибыли, руб": "plan_profit",   # ближайший прокси
    }
    FACT_MAP = {
        "Сумма заказов, руб": "fact_revenue",
        "Факт заказов, шт": "fact_orders",
        "Прибыль, руб": "fact_profit",
    }
    out = {}
    for code, indicators in (plan_fact or {}).items():
        row = {}
        for ind_name, vals in (indicators or {}).items():
            if not ind_name or not isinstance(vals, dict):
                continue
            totals = vals.get("totals")
            n = num(totals)
            if n is None:
                continue
            if ind_name in PLAN_MAP:
                row[PLAN_MAP[ind_name]] = n
            elif ind_name in FACT_MAP:
                row[FACT_MAP[ind_name]] = n
        # оставляем только SKU где есть хотя бы один настоящий план
        if row.get("plan_revenue") or row.get("plan_orders"):
            out[code] = row
    return out


def merge_plan_fact_into_may(plan_fact, sku_may):
    """План-Факт перебивает метрики мая."""
    INDICATOR_TO_FIELD = {
        "Сумма заказов, руб": "revenue",
        "Факт заказов, шт": "orders",
        "Прибыль, руб": "profit",
        "Маржинальность": "margin",
        "Рекламные расходы, руб": "ad_spend",
        "ДРР от рекламных продаж": "drr",
        "ДРР от продаж": "tacos",
        "Остаток": "stock",
        "Оборачиваемость, дней": "turnover",
        "Замороженный капитал": "frozen",
    }
    by_code = {row["code"]: row for row in sku_may}
    for code, indicators in plan_fact.items():
        target = by_code.get(code)
        if not target:
            continue
        for indicator, field in INDICATOR_TO_FIELD.items():
            v = num(indicators.get(indicator, {}).get("totals"))
            if v is not None:
                target[field] = v
    return list(by_code.values())


def compute_oos(plan_fact, prices):
    """Дни до OOS для TOP_LOCOMOTIVES_RISK."""
    last7 = [f"{d:02d}.05.2026" for d in range(13, 20)]
    out = {}
    for code in LOCO_RISK_CODES:
        stock = None
        if code in prices:
            stock = prices[code].get("stock_total")
        if stock is None and code in plan_fact:
            stock = num(plan_fact[code].get("Остаток", {}).get("totals"))
        avg7 = None
        if code in plan_fact:
            fakt = plan_fact[code].get("Факт заказов, шт", {})
            vals = [num(fakt.get(k)) for k in last7]
            vals = [v for v in vals if v is not None]
            if vals:
                avg7 = sum(vals) / len(vals)
        days_left = None
        if stock is not None and avg7 and avg7 > 0:
            days_left = stock / avg7
        out[code] = {
            "stock": stock,
            "avg7": round_n(avg7, 1),
            "days_left": round_n(days_left, 1),
        }
    return out


def compute_top_by_manager(sku_year, sku_monthly):
    """Подмерживаем менеджера к 04_SKU_итоги_года из последнего месяца,
    где артикул встречался. Возвращаем top-5 best/worst для каждого активного менеджера."""
    # 1. Карта code → manager (из самого свежего месяца, где артикул есть)
    code_to_manager = {}
    for month in sorted(sku_monthly.keys(), reverse=True):
        for row in sku_monthly[month]:
            code = row.get("code")
            mgr = row.get("manager")
            if code and mgr and code not in code_to_manager:
                code_to_manager[code] = mgr

    # 2. Раскидываем 04_итоги по менеджерам
    by_mgr = defaultdict(list)
    for r in sku_year:
        mgr = code_to_manager.get(r.get("code"))
        if mgr in ACTIVE_MANAGERS:
            by_mgr[mgr].append({
                "code": r.get("code"),
                "name": r.get("name"),
                "category": r.get("category"),
                "revenue": int(r["revenue"]) if r.get("revenue") is not None else None,
                "profit": int(r["profit"]) if r.get("profit") is not None else None,
                "margin": round_n(r.get("margin"), 1),
                "drr": round_n(r.get("drr"), 1),
            })

    # 3. Топ-5 по прибыли и худшие 5 по сумме «низкая маржа + высокий ДРР»
    out = {}
    for mgr, items in by_mgr.items():
        with_profit = [x for x in items if x["profit"] is not None]
        best5 = sorted(with_profit, key=lambda x: x["profit"], reverse=True)[:5]

        def worst_score(x):
            # чем меньше маржа и больше ДРР, тем хуже. Игнорим None.
            margin = x["margin"] if x["margin"] is not None else 0
            drr = x["drr"] if x["drr"] is not None else 0
            return drr - margin

        worst5 = sorted(items, key=worst_score, reverse=True)[:5]
        out[mgr] = {"best5": best5, "worst5": worst5}
    return out


def compute_stars(sku_monthly):
    """Срез по 4 артикулам-звёздам для каждого месяца."""
    out = {}
    for month, rows in sku_monthly.items():
        by_code = {r["code"]: r for r in rows}
        month_data = {}
        for code in STAR_CODES:
            if code in by_code:
                r = by_code[code]
                month_data[code] = {
                    "revenue": r.get("revenue"),
                    "profit": r.get("profit"),
                    "margin": r.get("margin"),
                    "tacos": r.get("tacos"),
                    "drr": r.get("drr"),
                    "stock": r.get("stock"),
                    "turnover": r.get("turnover"),
                    "manager": r.get("manager"),
                    "group": r.get("group"),
                }
        out[month] = month_data
    return out


def parse_daily_yesterday(sku_monthly, plan_fact=None):
    """v1.4: воронка (inputs/*воронка*.xlsx) — источник правды для дневных
    выручки/заказов. Прибыль/маржа — из MPSTATS profit_per_unit с пометкой
    is_stale если MPSTATS отстал >1 дня.

    Возвращает {last, prev} со снепшотами выручки/прибыли/маржи общими и
    по менеджерам, плюс staleness-флаги.
    """
    import openpyxl
    import re as _re
    from datetime import date as _d

    funnels = {}
    funnels_by_sku = {}  # {date: {code: orders_sku}}
    for path in sorted(p for p in INPUTS_DIR.glob("*.xlsx") if "воронка" in p.name.lower()):
        m = _re.search(r"с (\d+)-(\d+)-(\d+) по", path.name)
        if not m:
            continue
        date_csv = f"{m.group(1).zfill(2)}.{m.group(2).zfill(2)}.{m.group(3)}"
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb["Фильтры"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            wb.close()
            continue
        header = rows[1]
        values = rows[2]
        idx = {h: i for i, h in enumerate(header) if h}
        def get(name):
            i = idx.get(name)
            if i is None: return None
            v = values[i]
            return float(v) if isinstance(v, (int, float)) else None
        def get_alt(*names):
            for n in names:
                v = get(n)
                if v is not None:
                    return v
            return None
        funnels[date_csv] = {
            "revenue": get("Заказали на сумму, ₽"),
            "orders":  get_alt("Заказали, шт", "Заказали товаров, шт"),
            "buyouts_qty":    get("Выкупили, шт"),
            "buyouts_amount": get("Выкупили на сумму, ₽"),
            "avg_price":      get("Средняя цена, ₽"),
            "buyout_pct":     get("Процент выкупа"),
            "shows":   get("Показы"),
            "clicks":  get("Переходы в карточку"),
            "carts":   get("Положили в корзину"),
        }
        # лист Товары для разбивки по SKU
        try:
            ws2 = wb["Товары"]
            rr = list(ws2.iter_rows(values_only=True))
            if len(rr) >= 3:
                h2 = rr[1]
                idx2 = {h: i for i, h in enumerate(h2) if h}
                code_i = idx2.get("Артикул продавца")
                orders_i = idx2.get("Заказали, шт") or idx2.get("Заказали товаров, шт")
                rev_i = idx2.get("Заказали на сумму, ₽")
                day_data = {}
                for r in rr[2:]:
                    if code_i is None or r[code_i] is None:
                        continue
                    code = r[code_i]
                    o = r[orders_i] if orders_i is not None else None
                    rv = r[rev_i] if rev_i is not None else None
                    day_data[code] = {
                        "orders": float(o) if isinstance(o, (int, float)) else 0.0,
                        "revenue": float(rv) if isinstance(rv, (int, float)) else 0.0,
                    }
                funnels_by_sku[date_csv] = day_data
        except KeyError:
            pass
        wb.close()

    if not funnels:
        return None

    dates = sorted(funnels.keys(), key=_date_key)
    last_d = dates[-1]
    prev_d = dates[-2] if len(dates) >= 2 else None

    # profit_per_unit из MPSTATS (план-факт)
    profit_per_unit = {}
    if plan_fact:
        for code, indicators in plan_fact.items():
            pu = num((indicators.get("Прибыль на 1 выкуп, руб") or {}).get("totals"))
            if pu is not None:
                profit_per_unit[code] = pu

    # v2.0: дневная прибыль/выручка — из годового parquet (история) + текущий план-факт.
    # csv-year/Дневная_*.csv больше не используем (топ-200 артикулов было ограничением).
    daily_mpstats_profit = {}
    daily_mpstats_revenue = {}

    pivot = _load_history_pivot(["Прибыль, руб", "Сумма заказов, руб"])
    for code_map in (pivot.get("Прибыль, руб") or {}).values():
        for d, v in code_map.items():
            daily_mpstats_profit[d] = daily_mpstats_profit.get(d, 0) + v
    for code_map in (pivot.get("Сумма заказов, руб") or {}).values():
        for d, v in code_map.items():
            daily_mpstats_revenue[d] = daily_mpstats_revenue.get(d, 0) + v

    # 2. Свежие 3 дня из плана-факта — перебивают CSV
    if plan_fact:
        # сначала найдём какие дни есть в плане-факте с данными
        for code, indicators in plan_fact.items():
            prof_ind = indicators.get("Прибыль, руб") or {}
            rev_ind = indicators.get("Сумма заказов, руб") or {}
            for k, v in prof_ind.items():
                if k == "totals":
                    continue
                if isinstance(v, (int, float)):
                    # помечаем что для этого дня есть свежие данные → накопим заново
                    pass
        # двухэтапная сборка: ключи где есть план-факт → пересобрать с нуля
        pf_days_profit = set()
        pf_days_revenue = set()
        new_profit = {}
        new_revenue = {}
        for code, indicators in plan_fact.items():
            prof_ind = indicators.get("Прибыль, руб") or {}
            rev_ind = indicators.get("Сумма заказов, руб") or {}
            for k, v in prof_ind.items():
                if k == "totals":
                    continue
                if isinstance(v, (int, float)):
                    pf_days_profit.add(k)
                    new_profit[k] = new_profit.get(k, 0) + v
            for k, v in rev_ind.items():
                if k == "totals":
                    continue
                if isinstance(v, (int, float)):
                    pf_days_revenue.add(k)
                    new_revenue[k] = new_revenue.get(k, 0) + v
        # перекрываем csv данными плана-факта по тем дням, где они есть
        for k in pf_days_profit:
            daily_mpstats_profit[k] = new_profit[k]
        for k in pf_days_revenue:
            daily_mpstats_revenue[k] = new_revenue[k]

    def margin_for_day(date_str):
        r = daily_mpstats_revenue.get(date_str) or 0
        p = daily_mpstats_profit.get(date_str)
        if not r or p is None:
            return None
        return p / r * 100

    # mpstats_last_date — последний день с реальной разноской
    mpstats_last_date = None
    for k, v in daily_mpstats_revenue.items():
        if v and v > 0:
            if mpstats_last_date is None or _date_key(k) > _date_key(mpstats_last_date):
                mpstats_last_date = k
    today_str = _d.today().strftime("%d.%m.%Y")
    if mpstats_last_date:
        days_behind = (_d.today() - _d(_date_key(mpstats_last_date)[0],
                                        _date_key(mpstats_last_date)[1],
                                        _date_key(mpstats_last_date)[2])).days
    else:
        days_behind = None
    mpstats_is_stale = (days_behind is None) or (days_behind > 1)

    # карта code → manager (из последнего месячного среза, где есть колонка)
    code_to_manager = {}
    for month in sorted(sku_monthly.keys(), reverse=True):
        for r in sku_monthly[month]:
            c, m = r.get("code"), r.get("manager")
            if c and m and c not in code_to_manager:
                code_to_manager[c] = m

    # импорт оператора для аномалий
    from engine.operators import compute_anomaly_mad
    from engine.contexts import ANOMALY_RULES
    from datetime import timedelta as _td

    def history_for(date_str, metric_dict):
        """Возвращает значения metric за 7 закрытых дней до date_str
        (не сегодня, не сам день)."""
        if not date_str:
            return []
        all_dates_sorted = sorted(metric_dict.keys(), key=_date_key)
        target_key = _date_key(date_str)
        today_key = (_d.today().year, _d.today().month, _d.today().day)
        result = []
        for k in all_dates_sorted:
            kk = _date_key(k)
            if kk >= target_key:
                continue
            if kk >= today_key:
                continue
            result.append(metric_dict.get(k))
        # последние 7 закрытых
        return [v for v in result[-ANOMALY_RULES["window_days"]:] if v is not None]

    # Маржа дневная: считаем по всем закрытым дням
    margin_history_dict = {k: margin_for_day(k) for k in daily_mpstats_revenue
                           if margin_for_day(k) is not None}

    # v1.6: предрасчёт is_anomaly для ВСЕХ дней (нужно для make_display).
    # Для подмены отображения берём только sign_flip (реальный шум разноски),
    # mad_outlier не блокирует — это просто широкая дисперсия,
    # а не противоречие медиане.
    profit_anomaly_by_date = {}
    margin_anomaly_by_date = {}
    for d in daily_mpstats_profit.keys():
        h = history_for(d, daily_mpstats_profit)
        r = compute_anomaly_mad(daily_mpstats_profit[d], h,
                                mad_multiplier=ANOMALY_RULES["mad_multiplier"])
        profit_anomaly_by_date[d] = (r["reason"] == "sign_flip")
    for d in margin_history_dict.keys():
        h = history_for(d, margin_history_dict)
        r = compute_anomaly_mad(margin_history_dict[d], h,
                                mad_multiplier=ANOMALY_RULES["mad_multiplier"])
        margin_anomaly_by_date[d] = (r["reason"] == "sign_flip")

    def snapshot(date):
        if date is None:
            return None
        f = funnels.get(date, {})
        revenue = f.get("revenue")
        orders = f.get("orders")

        # v1.5: прибыль/маржа из дневной разноски MPSTATS
        profit_total = daily_mpstats_profit.get(date)
        margin_val = margin_for_day(date)
        profit_stale = mpstats_is_stale or profit_total is None

        # AnomalyDetect для отображения / истории
        prof_hist = history_for(date, daily_mpstats_profit)
        prof_anom = compute_anomaly_mad(profit_total, prof_hist,
                                        mad_multiplier=ANOMALY_RULES["mad_multiplier"])
        marg_hist = history_for(date, margin_history_dict)
        marg_anom = compute_anomaly_mad(margin_val, marg_hist,
                                        mad_multiplier=ANOMALY_RULES["mad_multiplier"])

        # v1.7: пометричные is_clean флаги — UI решает за метрику отдельно
        profit_is_clean = (profit_total is not None
                           and not mpstats_is_stale
                           and not profit_anomaly_by_date.get(date, False))
        margin_is_clean = (margin_val is not None
                           and not mpstats_is_stale
                           and not margin_anomaly_by_date.get(date, False))

        # v1.8: месячные агрегаты по чистым дням текущего месяца
        month_suffix = "." + ".".join(date.split(".")[1:]) if date else ""
        clean_profit_days = [d for d in daily_mpstats_profit
                             if d.endswith(month_suffix)
                             and not profit_anomaly_by_date.get(d, False)]
        profit_month_total = sum(daily_mpstats_profit[d] for d in clean_profit_days) if clean_profit_days else None
        profit_month_avg = (profit_month_total / len(clean_profit_days)) if clean_profit_days else None

        clean_margin_days = [d for d in margin_history_dict
                             if d.endswith(month_suffix)
                             and not margin_anomaly_by_date.get(d, False)]
        margin_month_avg = (sum(margin_history_dict[d] for d in clean_margin_days) / len(clean_margin_days)) if clean_margin_days else None

        # Разбивка по менеджеру: используем voronka.orders по SKU + profit_per_unit
        # как фолбэк (на уровне менеджера дневной MPSTATS-профит не очень информативен).
        by_sku = funnels_by_sku.get(date, {})
        by_mgr = {m: {"revenue": 0.0, "orders": 0.0, "profit": 0.0, "_n": 0} for m in ACTIVE_MANAGERS}
        for code, vals in by_sku.items():
            mgr = code_to_manager.get(code)
            if mgr not in by_mgr:
                continue
            by_mgr[mgr]["revenue"] += vals.get("revenue") or 0
            by_mgr[mgr]["orders"]  += vals.get("orders") or 0
            pu = profit_per_unit.get(code)
            o = vals.get("orders") or 0
            if pu is not None and o > 0:
                by_mgr[mgr]["profit"] += pu * o
                by_mgr[mgr]["_n"] += 1
        for mgr in by_mgr:
            r = by_mgr[mgr]
            r["margin"] = (r["profit"] / r["revenue"] * 100) if (r["revenue"] and r["_n"] >= 3) else None
            r["profit_is_stale"] = profit_stale or r["_n"] < 3
            r.pop("_n", None)

        return {
            "date": date,
            "revenue": round_n(revenue, 0) if revenue is not None else None,
            "orders": int(orders) if orders is not None else None,
            # v1.7: значение за тот же день + флаг is_clean. UI рисует "—" если !is_clean.
            "profit": round_n(profit_total, 0) if profit_total is not None else None,
            "profit_is_clean": profit_is_clean,
            "profit_is_stale": profit_stale,
            "profit_is_anomaly": prof_anom["is_anomaly"],
            "profit_anomaly_meta": prof_anom,
            "profit_raw": round_n(profit_total, 0) if profit_total is not None else None,
            "profit_month_total": round_n(profit_month_total, 0) if profit_month_total is not None else None,
            "profit_month_avg":   round_n(profit_month_avg, 0)   if profit_month_avg   is not None else None,
            "profit_month_clean_days": len(clean_profit_days),
            "margin": round_n(margin_val, 2) if margin_val is not None else None,
            "margin_is_clean": margin_is_clean,
            "margin_is_stale": profit_stale,
            "margin_is_anomaly": marg_anom["is_anomaly"],
            "margin_anomaly_meta": marg_anom,
            "margin_raw": round_n(margin_val, 2) if margin_val is not None else None,
            "margin_month_avg": round_n(margin_month_avg, 2) if margin_month_avg is not None else None,
            "margin_month_clean_days": len(clean_margin_days),
            "source": "funnel+mpstats",
            "by_manager": {m: {
                "revenue": round_n(v["revenue"], 0) if v["revenue"] else None,
                "orders": int(v["orders"]) if v["orders"] else None,
                "profit": round_n(v["profit"], 0) if v["profit"] else None,
                "margin": round_n(v["margin"], 1) if v["margin"] is not None else None,
                "profit_is_stale": v["profit_is_stale"],
            } for m, v in by_mgr.items()},
        }

    # v2.1: навигатор последних 7 дней
    last7 = dates[-7:]
    days = [snapshot(d) for d in last7]
    days = [d for d in days if d is not None]

    # v2.3 БЛОК 4: накопленная выручка по дням текущего месяца — для темпа на конкретную дату
    # ключ — день месяца (1..31), значение — суммарная выручка с 1 числа по этот день включительно.
    month_cumulative = {}
    cur_month_suffix = "." + ".".join(last_d.split(".")[1:]) if last_d else ""
    days_in_cur_month = sorted([d for d in funnels if d.endswith(cur_month_suffix)], key=_date_key)
    running = 0.0
    for d in days_in_cur_month:
        rv = (funnels.get(d) or {}).get("revenue") or 0
        running += rv
        try:
            day_n = int(d.split(".")[0])
            month_cumulative[day_n] = round_n(running, 0)
        except (ValueError, IndexError):
            pass

    return {
        "last": snapshot(last_d),
        "prev": snapshot(prev_d),
        "days": days,
        "month_cumulative_revenue": month_cumulative,
        "mpstats_meta": {
            "last_known_date": mpstats_last_date,
            "today": today_str,
            "days_behind": days_behind,
            "is_stale": mpstats_is_stale,
        },
    }


METRIC_TO_INDICATOR = {
    "orders":  "Факт заказов, шт",
    "price":   "Цена после СПП",
    "spp":     "СПП",
    "stock":   "Остаток",
    "profit":  "Прибыль, руб",
    # v2.1: ДРР и Конверсия для строк changes в карточках
    "drr":     "ДРР от рекламных продаж",
    "conv":    "Конверсия в заказ",
}


_INT_METRICS = {"orders", "stock"}
_PRECISE_METRICS = {"drr", "conv", "spp"}  # 2 знака


def _compact_value(metric, v):
    if v is None:
        return None
    if metric in _INT_METRICS:
        return int(round(v))
    if metric in _PRECISE_METRICS:
        return round(v, 2)
    return round(v, 1)


def parse_daily_history(sku_monthly, current_month, window_days=30):
    """v2.2: дневная история 30 дней (нужно для переключателя «30 дней» в модалке).
    Покрытие — все SKU из годового архива.
    Возвращает sku_history[code] = {dates, orders, price, spp, stock, profit, drr, conv}.
    """
    pivot = _load_history_pivot(list(METRIC_TO_INDICATOR.values()))
    if not pivot:
        return {}

    all_dates = set()
    for ind in METRIC_TO_INDICATOR.values():
        for code_map in (pivot.get(ind) or {}).values():
            all_dates.update(code_map.keys())
    dates_n = sorted(all_dates, key=_date_key)[-window_days:]

    all_codes = set()
    for ind in METRIC_TO_INDICATOR.values():
        all_codes.update((pivot.get(ind) or {}).keys())
    all_codes |= STAR_CODES | LOCO_RISK_CODES

    out = {}
    for code in all_codes:
        per = {"dates": dates_n}
        any_data = False
        for metric, ind in METRIC_TO_INDICATOR.items():
            code_map = (pivot.get(ind) or {}).get(code, {})
            arr = [_compact_value(metric, code_map.get(d)) for d in dates_n]
            per[metric] = arr
            if any(v is not None for v in arr):
                any_data = True
        if any_data:
            out[code] = per
    return out


def parse_monthly_history(window_months=12):
    """v2.0: помесячные агрегаты для графика «12 месяцев» в модалке SKU."""
    pivot = _load_history_pivot(["Факт заказов, шт", "Сумма заказов, руб", "Прибыль, руб"])
    if not pivot:
        return {}

    def month_of(date_str):
        try:
            d, m, y = date_str.split(".")
            return f"{y}-{m}"
        except Exception:
            return None

    all_months = set()
    for ind in ("Факт заказов, шт", "Сумма заказов, руб", "Прибыль, руб"):
        for code_map in (pivot.get(ind) or {}).values():
            for d in code_map.keys():
                m = month_of(d)
                if m:
                    all_months.add(m)
    months_sorted = sorted(all_months)[-window_months:]

    all_codes = set()
    for ind in ("Факт заказов, шт", "Сумма заказов, руб", "Прибыль, руб"):
        all_codes.update((pivot.get(ind) or {}).keys())

    out = {}
    for code in all_codes:
        per = {"months": months_sorted, "orders": [], "revenue": [], "profit": []}
        any_data = False
        for m in months_sorted:
            for metric, ind in (("orders", "Факт заказов, шт"),
                                ("revenue", "Сумма заказов, руб"),
                                ("profit", "Прибыль, руб")):
                code_map = (pivot.get(ind) or {}).get(code, {})
                total = sum(v for d, v in code_map.items() if month_of(d) == m and v is not None)
                # сжатие: orders как int, revenue/profit в ТЫСЯЧАХ для экономии места
                if metric == "orders":
                    per[metric].append(int(round(total)) if total else 0)
                else:
                    per[metric].append(int(round(total / 1000)) if total else 0)
                if total:
                    any_data = True
        if any_data:
            out[code] = per
    return out


def read_ai_summary():
    """Читает data/ai_summary.json если есть, иначе заглушку."""
    if AI_SUMMARY_FILE.exists():
        try:
            return json.loads(AI_SUMMARY_FILE.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"  ⚠ ai_summary.json повреждён: {e}")
    return {
        "company_summary": "AI-сводка не настроена или ещё не сгенерирована. Запусти `python3 update.py` — Gemini проанализирует данные за вчера и сгенерирует сводку.",
        "manager_blocks": {"Виктория": "", "Настя": "", "Владимир": ""},
        "tasks": [],
        "generated_at": None,
    }


def read_daily_summary():
    if DAILY_SUMMARY_MD.exists():
        return DAILY_SUMMARY_MD.read_text(encoding="utf-8")
    return ("# Сводка дня\n\n"
            "_Файл `data/daily_summary.md` не найден. "
            "Создайте его — он будет показан здесь._\n")


def latest_mtime():
    paths = list(CSV_DIR.glob("*.csv")) + list(INPUTS_DIR.glob("*.xlsx"))
    if not paths:
        return datetime.now().isoformat(timespec="seconds")
    return datetime.fromtimestamp(max(p.stat().st_mtime for p in paths)).isoformat(timespec="seconds")


INT_FIELDS = {"revenue", "orders", "profit", "ad_spend", "frozen", "stock"}
DEC1_FIELDS = {"margin", "drr", "tacos", "turnover"}


def compact_sku_row(row):
    """Урезаем числа до int/1 знака — экономия места в months/*.json."""
    out = {}
    for k, v in row.items():
        if k in INT_FIELDS and isinstance(v, (int, float)):
            out[k] = int(round(v))
        elif k in DEC1_FIELDS and isinstance(v, (int, float)):
            out[k] = round(v, 1)
        else:
            out[k] = v
    return out


def emit(common_data, months_data):
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    MONTHS_DIR.mkdir(parents=True, exist_ok=True)

    compact_months = {m: [compact_sku_row(r) for r in rows] for m, rows in months_data.items()}

    with open(DATA_JSON, "w", encoding="utf-8") as f:
        json.dump(common_data, f, ensure_ascii=False, separators=(",", ":"))

    for month, rows in compact_months.items():
        with open(MONTHS_DIR / f"{month}.json", "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False, separators=(",", ":"))

    if INDEX_HTML.exists():
        html = INDEX_HTML.read_text(encoding="utf-8")
        common_json = json.dumps(common_data, ensure_ascii=False, separators=(",", ":"))
        months_json = json.dumps(compact_months, ensure_ascii=False, separators=(",", ":"))

        # 1. Первый прогон: маркеры присутствуют → простая замена.
        # 2. Повторные прогоны: маркеров уже нет, заменяем по anchor-границам.
        if "/* APP_DATA_PLACEHOLDER */" in html:
            html = html.replace("/* APP_DATA_PLACEHOLDER */", common_json)
            html = html.replace("/* MONTHS_DATA_PLACEHOLDER */", months_json)
        else:
            import re
            # APP_DATA — от 'const APP_DATA = ' до ';\nconst MONTHS_DATA = '
            html = re.sub(
                r"const APP_DATA = .+?;\nconst MONTHS_DATA = ",
                "const APP_DATA = " + common_json.replace("\\", "\\\\") + ";\nconst MONTHS_DATA = ",
                html, count=1, flags=re.DOTALL,
            )
            # MONTHS_DATA — от 'const MONTHS_DATA = ' до ';\n' до конца JS-блока.
            # Ищем строку 'const MONTHS_DATA = ...;\n' (до перевода + следующей строки кода).
            html = re.sub(
                r"const MONTHS_DATA = .+?;\n",
                "const MONTHS_DATA = " + months_json.replace("\\", "\\\\") + ";\n",
                html, count=1, flags=re.DOTALL,
            )
        INDEX_HTML.write_text(html, encoding="utf-8")


def main():
    print("[1/8] CSV-агрегаты...")
    summary = parse_summary_csv()
    managers = parse_managers_csv()
    categories = parse_categories_csv()
    sku_year = parse_sku_year_csv()

    print("[2/8] Месячные SKU (12 файлов)...")
    sku_monthly = parse_sku_monthly_csv()

    print("[3/8] inputs/План-Факт...")
    plan_fact, pf_name = parse_plan_fact_xlsx()
    print(f"        источник: {pf_name}, артикулов: {len(plan_fact)}")

    print("[4/8] inputs/Цены...")
    prices, prices_name = parse_prices_xlsx()
    print(f"        источник: {prices_name}, артикулов: {len(prices)}")

    print("[5/8] inputs/Воронки...")
    funnels = parse_funnels_xlsx()
    print(f"        дней с воронками: {len(funnels['overall'])} ({sorted(funnels['overall'].keys())})")

    print("[6/8] Перебиваем май свежим планом-фактом...")
    if "2026-05" in sku_monthly and plan_fact:
        sku_monthly["2026-05"] = merge_plan_fact_into_may(plan_fact, sku_monthly["2026-05"])

    print("[7/8] OOS + срезы звёзд + топы по менеджерам...")
    oos = compute_oos(plan_fact, prices)
    stars = compute_stars(sku_monthly)
    top_by_manager = compute_top_by_manager(sku_year, sku_monthly)

    print("[7b] дневные данные из воронки + MPSTATS staleness...")
    yesterday = parse_daily_yesterday(sku_monthly, plan_fact=plan_fact)
    if yesterday and yesterday.get("last"):
        print(f"        последний день в данных: {yesterday['last']['date']}")

    print("[7c] 30-дневная история из history.parquet...")
    sku_history = parse_daily_history(sku_monthly, DEFAULT_MONTH)
    print(f"        SKU в дневной истории: {len(sku_history)}")

    print("[7c.2] 12-месячная история для модалок SKU...")
    sku_history_monthly_full = parse_monthly_history(window_months=12)
    # Ужимаем: оставляем только SKU из текущего месячного среза (1003), +звёзды/локо
    active_codes = {r["code"] for r in sku_monthly.get(DEFAULT_MONTH, []) if r.get("code")}
    active_codes |= STAR_CODES | LOCO_RISK_CODES
    sku_history_monthly = {c: v for c, v in sku_history_monthly_full.items() if c in active_codes}
    print(f"        SKU в месячной истории (после фильтрации active): {len(sku_history_monthly)}")

    # v1.9: coverage-лог по текущему месяцу
    month_rows = sku_monthly.get(DEFAULT_MONTH, []) or []
    total = len(month_rows)
    with_basics = 0
    with_7d = 0
    with_14d = 0
    no_history_zero_orders = 0
    no_history_new_sku = 0
    for r in month_rows:
        if r.get("revenue") is not None and r.get("stock") is not None:
            with_basics += 1
        hist = sku_history.get(r.get("code"))
        if hist:
            orders_arr = [o for o in (hist.get("orders") or []) if o is not None]
            n_days = len(orders_arr)
            if n_days >= 7:
                with_7d += 1
            if n_days >= 14:
                with_14d += 1
        else:
            if (r.get("orders") or 0) == 0:
                no_history_zero_orders += 1
            else:
                no_history_new_sku += 1
    def _pct(x, n): return f"{x*100/n:.0f}%" if n else "—"
    print(f"        COVERAGE: всего {total}, с базовыми полями {with_basics} ({_pct(with_basics, total)})")
    print(f"                  с 7д динамикой {with_7d} ({_pct(with_7d, total)})")
    print(f"                  с 14д динамикой {with_14d} ({_pct(with_14d, total)})")
    print(f"                  без истории: 0-заказов {no_history_zero_orders}, новые/прочие {no_history_new_sku}")

    print("[7d] ai_summary.json...")
    ai_summary = read_ai_summary()

    print("[8/8] daily_summary.md и финальная сборка...")
    summary_md = read_daily_summary()

    sku_plans = compute_sku_plans(plan_fact)
    print(f"        SKU c планами в xlsx: {len(sku_plans)}")

    month_plans = compute_month_plans(DEFAULT_MONTH)
    if month_plans:
        mp = month_plans[DEFAULT_MONTH]
        print(f"        month_plans {DEFAULT_MONTH}: rev={mp.get('revenue')}, "
              f"sales={mp.get('sales_qty')}, profit={mp.get('profit')}, "
              f"margin={mp.get('margin_pct')}%")

    month_buyouts = compute_month_buyouts(DEFAULT_MONTH)
    if month_buyouts:
        mb = month_buyouts[DEFAULT_MONTH]
        print(f"        month_buyouts {DEFAULT_MONTH}: orders={mb.get('orders_amount')}, "
              f"buyouts={mb.get('buyouts_amount')}, pct={mb.get('buyout_pct')}%")

    common = {
        "meta": {
            "last_updated": latest_mtime(),
            "current_month": DEFAULT_MONTH,
            "incomplete_months": sorted(INCOMPLETE_MONTHS),
            "no_managers_months": sorted(NO_MANAGERS_MONTHS),
            "available_months": [r["month"] for r in summary],
            "sources": {
                "plan_fact": pf_name,
                "prices": prices_name,
                "funnels_days": sorted(funnels["overall"].keys()),
            },
        },
        "monthly_totals": summary,
        "manager_monthly": managers,
        "category_monthly": categories,
        "top_by_manager": top_by_manager,
        "stars_monthly": stars,
        "oos_days": oos,
        "prices": prices,
        "funnels": funnels,
        "yesterday": yesterday,
        "sku_history": sku_history,
        "sku_history_monthly": sku_history_monthly,
        "sku_plans": sku_plans,
        "month_plans": month_plans,
        "month_buyouts": month_buyouts,
        "ai_summary": ai_summary,
        "daily_summary_md": summary_md,
    }

    emit(common, sku_monthly)

    total = DATA_JSON.stat().st_size
    months_total = sum(p.stat().st_size for p in MONTHS_DIR.glob("*.json"))
    print()
    print(f"data.json: {total/1024:.1f} КБ")
    print(f"data/processed/months/*.json: {months_total/1024:.1f} КБ "
          f"({len(list(MONTHS_DIR.glob('*.json')))} файлов)")
    print(f"Суммарно: {(total+months_total)/1024/1024:.2f} МБ")


if __name__ == "__main__":
    main()
